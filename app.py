import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import streamlit.components.v1 as components

# Colores
COLOR_VERDE = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
COLOR_MORADO = PatternFill(start_color="800080", end_color="800080", fill_type="solid")

st.set_page_config(page_title="Inventario Biblioteca", layout="centered")
st.title("📚 Inventario Biblioteca")
st.write("Escanea códigos automáticamente con la cámara y actualiza el inventario.")

# Ruta del Excel por defecto
DEFAULT_EXCEL = "inventario.xlsx"

# Si no existe, creamos uno básico
import os
from openpyxl import Workbook
if not os.path.exists(DEFAULT_EXCEL):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Inventario"
    sheet.append(["Codigo", "Titulo", "Autor"])
    sheet.append(["12345", "El Quijote", "Cervantes"])
    sheet.append(["67890", "Cien Años de Soledad", "García Márquez"])
    wb.save(DEFAULT_EXCEL)

# Subir archivo Excel opcional
uploaded_file = st.file_uploader("Sube tu archivo Excel del inventario (opcional)", type=["xlsx"])
if uploaded_file:
    excel_path = "inventario.xlsx"
    with open(excel_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
else:
    excel_path = DEFAULT_EXCEL

# Cargar Excel
wb = load_workbook(excel_path)
sheet = wb.active
df = pd.read_excel(excel_path)

# Detectar columna de códigos
codigo_columna = None
for col in df.columns:
    if "codigo" in col.lower():
        codigo_columna = col
        break

if not codigo_columna:
    st.error("No se encontró ninguna columna que contenga 'codigo'.")
else:
    codigo_a_fila = {str(row[codigo_columna]): idx+2 for idx, row in df.iterrows()}

    st.subheader("Escanea el código automáticamente")
    # Componente QuaggaJS
    components.html(
        """
        <div id="interactive" class="viewport" style="width:100%; height:300px;"></div>
        https://cdnjs.cloudflare.com/ajax/libs/quagga/0.12.1/quagga.min.js</script>
        <script>
        const config = {
            inputStream: {
                type: "LiveStream",
                constraints: {
                    facingMode: "environment" // cámara trasera
                },
                target: document.querySelector('#interactive')
            },
            decoder: {
                readers: ["code_128_reader", "ean_reader", "ean_8_reader", "code_39_reader", "upc_reader"]
            }
        };

        Quagga.init(config, function(err) {
            if (err) {
                console.log(err);
                return;
            }
            Quagga.start();
        });

        Quagga.onDetected(function(result) {
            const code = result.codeResult.code;
            window.parent.postMessage({type: 'barcode', value: code}, '*');
        });
        </script>
        """,
        height=350,
    )

    # Capturar código detectado con streamlit-javascript
    from streamlit_javascript import st_javascript
    barcode = st_javascript("await new Promise(resolve => {window.addEventListener('message', e => {if(e.data.type==='barcode'){resolve(e.data.value)}})})")
    if barcode:
        st.success(f"Código detectado automáticamente: {barcode}")

    # Campo manual opcional
    codigo_manual = st.text_input("Ingresa el código manualmente (opcional)", value=barcode if barcode else "")

    if st.button("Actualizar Inventario"):
        if codigo_manual.strip() != "":
            codigo = codigo_manual.strip()
            if codigo in codigo_a_fila:
                fila = codigo_a_fila[codigo]
                celda = f"A{fila}"
                sheet[celda].fill = COLOR_VERDE
                sheet[celda].font = Font(bold=True)
                st.success(f"Código {codigo} encontrado y marcado en verde.")
            else:
                nueva_fila = sheet.max_row + 1
                sheet[f"A{nueva_fila}"] = codigo
                sheet[f"A{nueva_fila}"].fill = COLOR_MORADO
                sheet[f"A{nueva_fila}"].font = Font(bold=True)
                st.warning(f"Código {codigo} agregado como nuevo y marcado en morado.")

            wb.save(excel_path)
        else:
            st.error("Por favor, escanea o ingresa el código manualmente.")

    st.subheader("Inventario actualizado")
    st.dataframe(pd.read_excel(excel_path))

    with open(excel_path, "rb") as f:
        st.download_button("Descargar Excel actualizado", f, file_name="inventario_actualizado.xlsx")
